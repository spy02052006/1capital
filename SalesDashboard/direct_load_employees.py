#!/usr/bin/env python3
"""
Direct data loader for Employee data from Excel file
Uses SQLite direct insertion instead of pandas/data_pipeline
"""

import os
import sys
import sqlite3
import csv
from pathlib import Path

# Find the Excel file and convert to CSV
data_dir = Path('/var/www/SalesDashboardProject/data_files/Employee_dim')
excel_file = data_dir / 'Wirecode_wise_RMdetails.xlsx'

print(f"Data directory: {data_dir}")
print(f"Excel file exists: {excel_file.exists()}")

if excel_file.exists():
    # Use xlrd/openpyxl to read Excel if available
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(str(excel_file))
        ws = wb.active
        
        print(f"Excel loaded, active sheet: {ws.title}")
        
        # Connect to database
        db_path = '/var/www/SalesDashboardProject/SalesDashboard/db.sqlite3'
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Clear existing data
        cursor.execute("DELETE FROM employee_dimension")
        print("Cleared existing employee records")
        
        # Read header row
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value.lower().strip()] = col_idx
        
        print(f"Headers found: {list(headers.keys())}")
        
        # Load data rows
        count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Extract values by position (Column A=0, B=1, D=3, E=4)
                emp_id = int(row[0]) if row[0] and str(row[0]).strip().isdigit() else None
                rm_name = str(row[1]).strip() if row[1] else None
                # pan = str(row[2]).strip() if row[2] else None  # Column C
                manager_id = int(row[3]) if row[3] and str(row[3]).strip().isdigit() else None  # Column D
                wire_code = str(row[4]).strip() if row[4] else f'EMP_{emp_id}'  # Column E
                
                if not emp_id or not rm_name:
                    print(f"Row {row_idx}: Skipped (missing ID or NAME)")
                    continue
                
                # Insert into database
                cursor.execute("""
                    INSERT INTO employee_dimension 
                    (id, wire_code, rm_name, manager_id, is_active, created_at, updated_at)
                    VALUES (?, ?, ?, ?, 1, datetime('now'), datetime('now'))
                """, (emp_id, wire_code, rm_name, manager_id))
                
                count += 1
                print(f"Row {row_idx}: ID={emp_id}, Name={rm_name}, Manager={manager_id}")
            
            except Exception as e:
                print(f"Row {row_idx}: Error - {e}")
                continue
        
        conn.commit()
        
        # Verify
        cursor.execute("SELECT COUNT(*) FROM employee_dimension")
        total = cursor.fetchone()[0]
        print(f"\nTotal employees loaded: {total}")
        
        # Show sample
        cursor.execute("SELECT id, rm_name, manager_id, wire_code FROM employee_dimension LIMIT 3")
        rows = cursor.fetchall()
        print("Sample records:")
        for emp_id, rm_name, manager_id, wire_code in rows:
            print(f"  ID={emp_id}, Name={rm_name}, Manager={manager_id}, Wire={wire_code}")
        
        conn.close()
        print("\nEmployee reload completed!")
    
    except ImportError as e:
        print(f"openpyxl not available: {e}")
else:
    print(f"Excel file not found: {excel_file}")
